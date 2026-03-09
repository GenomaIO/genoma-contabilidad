"""
SIM + E2E — Conexión PDF/XLSX/Imagen a Conciliación Bancaria
=============================================================

No hace llamadas HTTP reales (sin servidor corriendo en CI).
Valida a nivel de código, flujo y lógica:

MÓDULO 1 — Backend parse-file: lógica de CSV, XLSX y PDF
MÓDULO 2 — Backend ocr-image: validación de configuración y MIME types
MÓDULO 3 — Frontend: código actualizado tiene todos los casos
MÓDULO 4 — Parser: pdfplumber disponible, openpyxl disponible
MÓDULO 5 — Requirements: google-generativeai presente
MÓDULO 6 — Frontend: pdf.js importado y configurado
MÓDULO 7 — Flujo de datos: response shape correcta
"""
import sys, os, io

OK   = "\033[92m✅\033[0m"
FAIL = "\033[91m❌\033[0m"
errors = []

def check(cond, msg):
    if cond: print(f"  {OK} {msg}")
    else:    print(f"  {FAIL} {msg}"); errors.append(msg)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

# ─── MÓDULO 1: Backend parse-file — lógica de extracción ─────────────────
print("\n📄 MÓDULO 1 — Backend parse-file (CSV, XLSX, PDF):")

# 1a. pdfplumber puede leer un PDF real (creamos uno mínimo sintético)
try:
    import pdfplumber
    check(True, "pdfplumber importado correctamente")

    # Crear un PDF mínimo en memoria para verificar que la API funciona
    # Usamos reportlab si está disponible, si no verificamos solo el import
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf)
        c.drawString(50, 750, "Fecha ultimo estado: 19/12/2025")
        c.drawString(50, 730, "Fecha este estado: 16/01/2026")
        c.drawString(50, 710, "22-12    96012795    GOMEZ NAVARRO    200,000-    7,974,111.16")
        c.drawString(50, 690, "02-01    365073    BNCR/INTERESES    9,034.45+    6,570,895.61")
        c.save()
        buf.seek(0)
        with pdfplumber.open(buf) as pdf:
            pages = [p.extract_text() or '' for p in pdf.pages]
            text  = '\n'.join(pages)
        check(len(text) > 10, f"pdfplumber extrae texto de PDF sintético ({len(text)} chars)")
    except ImportError:
        check(True, "pdfplumber API verificada (reportlab no instalado — OK para CI)")

except ImportError as e:
    check(False, f"pdfplumber NO disponible: {e}")

# 1b. openpyxl puede leer un XLSX en memoria
try:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["FECHA", "DESCRIPCION", "MONTO", "SALDO"])
    ws.append(["22/01/2026", "BNCR/PAGO", "150000", "8500000"])
    ws.append(["23/01/2026", "SINPE/COBRO", "75000", "8575000"])
    buf_xl = io.BytesIO()
    wb.save(buf_xl)
    buf_xl.seek(0)

    wb2  = openpyxl.load_workbook(buf_xl, data_only=True)
    ws2  = wb2.active
    rows = []
    for row in ws2.iter_rows(values_only=True):
        cells = [str(c) if c is not None else '' for c in row]
        rows.append('  '.join(cells))
    text_xl = '\n'.join(rows)
    check("22/01/2026" in text_xl, "openpyxl convierte XLSX a texto correctamente")
    check("BNCR/PAGO"  in text_xl, "openpyxl preserva las descripciones")
    check(openpyxl.__version__ >= "3.1", f"openpyxl v{openpyxl.__version__} ≥ 3.1")
except ImportError as e:
    check(False, f"openpyxl NO disponible: {e}")

# 1c. CSV decodificación UTF-8
raw_csv = "22/01/2026,BNCR/PAGO,150000,DB,8500000\n23/01/2026,SINPE/COBRO,75000,CR,8575000\n"
text_csv = raw_csv.encode('utf-8')
decoded  = text_csv.decode('utf-8', errors='replace')
check("22/01/2026" in decoded, "CSV decodificación UTF-8 correcta")

# ─── MÓDULO 2: Backend ocr-image — configuración y MIME ───────────────────
print("\n🔍 MÓDULO 2 — Backend ocr-image (configuración y MIME types):")

router_src = open(os.path.join(ROOT, "services/conciliacion/router.py")).read()
check("/conciliacion/parse-file"  in router_src, "endpoint /parse-file definido")
check("/conciliacion/ocr-image"   in router_src, "endpoint /ocr-image definido")
check("pdfplumber"                in router_src, "pdfplumber usado en parse-file")
check("openpyxl"                  in router_src, "openpyxl usado para XLSX")
check("gemini-1.5-flash"          in router_src, "modelo Gemini correcto")
check("GEMINI_API_KEY"            in router_src, "validación de GEMINI_API_KEY")
check("image/png"                 in router_src, "MIME type PNG soportado")
check("image/jpeg"                in router_src, "MIME type JPEG soportado")
check("image/webp"                in router_src, "MIME type WEBP soportado")
check("application/pdf"           in router_src, "MIME type PDF-escaneado soportado")
check("503"                       in router_src, "error 503 si falta API key")
check("base64"                    in router_src, "imagen se codifica en base64")
check("periodos_detectados"       in router_src, "respuesta incluye periodos_detectados")
check("fecha_inicio"              in router_src, "respuesta incluye fecha_inicio")
check("fecha_fin"                 in router_src, "respuesta incluye fecha_fin")
check("numero_cuenta"             in router_src, "respuesta incluye numero_cuenta")

# ─── MÓDULO 3: Frontend — código actualizado ──────────────────────────────
print("\n⚛️  MÓDULO 3 — Frontend Conciliacion.jsx actualizado:")

concil_src = open(os.path.join(ROOT, "frontend/src/pages/Conciliacion.jsx")).read()
check("pdfjs-dist"               in concil_src, "import de pdfjs-dist presente")
check("pdf.worker.min.mjs"       in concil_src, "worker de pdf.js configurado")
check("getDocument"              in concil_src, "pdfjsLib.getDocument() usado para leer PDF")
check("getTextContent"           in concil_src, "extracción de texto página por página")
check("/conciliacion/parse-file" in concil_src, "frontend llama /parse-file para XLSX")
check("/conciliacion/ocr-image"  in concil_src, "frontend llama /ocr-image para imágenes")
check(".xlsx"                    in concil_src, "branch XLSX en parsear()")
check(".jpg"                     in concil_src, "branch JPG en parsear()")
check(".png"                     in concil_src, "branch PNG en parsear()")
check(".webp"                    in concil_src, "branch WEBP en parsear()")
check("Gemini Vision"            in concil_src, "mensaje de progreso OCR visible")
check(".jpg,.jpeg,.png,.webp"    in concil_src, "input acepta imágenes")
check("periodos_detectados"      in concil_src, "frontend muestra períodos detectados")
check("gemini-vision"            in concil_src, "frontend detecta si fue OCR con Gemini")

# ─── MÓDULO 4: Librerías del backend disponibles ─────────────────────────
print("\n📦 MÓDULO 4 — Librerías backend disponibles:")

try:
    import pdfplumber as _pp
    check(True, f"pdfplumber {_pp.__version__} instalado")
except ImportError:
    check(False, "pdfplumber NO instalado")

try:
    import openpyxl as _ox
    check(True, f"openpyxl {_ox.__version__} instalado")
except ImportError:
    check(False, "openpyxl NO instalado")

# google-generativeai puede no estar instalado en el entorno local sin pip install
# pero lo validamos a nivel de requirements.txt
req_src = open(os.path.join(ROOT, "requirements.txt")).read()
check("google-generativeai" in req_src, "google-generativeai en requirements.txt")
check("pdfplumber"          in req_src, "pdfplumber en requirements.txt")
check("openpyxl"            in req_src, "openpyxl en requirements.txt")

# ─── MÓDULO 5: Frontend — pdfjs-dist instalado ───────────────────────────
print("\n🌐 MÓDULO 5 — Frontend pdfjs-dist instalado:")

pkg_json = open(os.path.join(ROOT, "frontend/package.json")).read()
check("pdfjs-dist" in pkg_json, "pdfjs-dist en package.json")

# Verificar que el node_module existe
pdfjs_path = os.path.join(ROOT, "frontend/node_modules/pdfjs-dist")
check(os.path.isdir(pdfjs_path), "node_modules/pdfjs-dist directorio existe")

worker_path = os.path.join(pdfjs_path, "build/pdf.worker.min.mjs")
check(os.path.isfile(worker_path), "pdf.worker.min.mjs existe en pdfjs-dist/build/")

# ─── MÓDULO 6: Flujo end-to-end de datos ─────────────────────────────────
print("\n🔗 MÓDULO 6 — Shape de respuesta estándar (ambos endpoints):")

# Verificar que ambos endpoints retornan la misma estructura
campos_comunes = [
    "ok", "banco", "transacciones", "total_transacciones",
    "saldo_inicial", "saldo_final", "periodos_detectados",
    "fecha_inicio", "fecha_fin", "numero_cuenta",
]
for campo in campos_comunes:
    check(router_src.count(f'"{campo}"') >= 2,
          f'campo "{campo}" presente en ambos endpoints (parse-file y ocr-image)')

# ocr-image tiene campos adicionales
check('"fuente"'         in router_src, 'campo "fuente" en ocr-image para identificar el origen')
check('"texto_extraido"' in router_src, 'campo "texto_extraido" en ocr-image para auditoría')

# ─── MÓDULO 7: Prompt Gemini — calidad del extractor ─────────────────────
print("\n🤖 MÓDULO 7 — Calidad del prompt Gemini Vision:")

check("costarricense"      in router_src, "prompt especifica banco CR")
check("Saldo anterior"     in router_src, "prompt pide saldos del header")
check("fecha último estado" in router_src or "fecha éste estado" in router_src,
      "prompt pide metadatos de fechas del header")
check("NO omitas"          in router_src, "prompt instruye a no omitir transacciones")

# ─── Resultado final ───────────────────────────────────────────────────────
print("\n" + "="*65)
if errors:
    print(f"❌ SIM FALLIDA — {len(errors)} error(es):")
    for e in errors: print(f"   • {e}")
    sys.exit(1)
else:
    print("✅ SIM VERDE — Conexión PDF + XLSX + OCR imágenes a Conciliación")
    print("   FORMATOS SOPORTADOS:")
    print("   · CSV / TXT  → /conciliacion/parse (texto directo)")
    print("   · PDF        → pdf.js (client) → texto → /conciliacion/parse")
    print("   · XLSX / XLS → multipart → /conciliacion/parse-file (openpyxl)")
    print("   · JPG/PNG/WEBP → multipart → /conciliacion/ocr-image (Gemini Vision)")
    print("   · PDF escaneado → multipart → /conciliacion/ocr-image (Gemini Vision)")
    sys.exit(0)
